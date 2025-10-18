import os
import sys
import json
import time
import logging
import torch
import random
import numpy as np
import pandas as pd
import pickle
import traceback
import torch.distributed as dist
import torch.nn.functional as F
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict
from torch.utils.data import DataLoader, Dataset
from torch.utils.data.distributed import DistributedSampler
from torch.nn.parallel import DistributedDataParallel as DDP

# Ensure project root is in path
sys.path.append("../")
from model.kronos import KronosTokenizer, Kronos, auto_regressive_inference
from common_data_processor import FinancialDataset

try:
    import comet_ml
except ImportError:
    comet_ml = None

# Setup logger
logger = logging.getLogger('KronosPipeline')


def get_shanghai_time():
    """
    获取上海时区的当前时间
    
    Returns:
        datetime: 上海时区的当前时间
    """
    try:
        import pytz
        shanghai_tz = pytz.timezone('Asia/Shanghai')
        return datetime.now(shanghai_tz)
    except ImportError:
        # 如果没有pytz，使用UTC+8
        from datetime import timezone
        shanghai_tz = timezone(timedelta(hours=8))
        return datetime.now(shanghai_tz)


def get_future_business_days(start_date, num_days):
    """
    生成未来N个工作日（跳过周末）
    
    Args:
        start_date: 起始日期
        num_days: 需要生成的工作日数量
        
    Returns:
        list: 未来N个工作日的日期列表
    """
    business_days = []
    current_date = start_date
    
    while len(business_days) < num_days:
        current_date = current_date + timedelta(days=1)
        # 0-4 代表周一到周五（工作日）
        if current_date.weekday() < 5:
            business_days.append(current_date)
    
    return business_days


def setup_logging(log_level=logging.INFO, log_dir='./logs'):
    """
    设置日志配置，使用上海时区，保存到文件（带日期后缀）
    
    Args:
        log_level: 日志级别
        log_dir: 日志保存目录
    """
    import pytz
    
    # 设置上海时区
    shanghai_tz = pytz.timezone('Asia/Shanghai')
    
    # 创建日志目录
    os.makedirs(log_dir, exist_ok=True)
    
    # 生成带日期后缀的日志文件名（使用上海时间）
    current_time = datetime.now(shanghai_tz)
    log_filename = current_time.strftime('training_%Y%m%d.log')
    log_filepath = os.path.join(log_dir, log_filename)
    
    # 自定义日志格式化器，使用上海时间
    class ShanghaiFormatter(logging.Formatter):
        def formatTime(self, record, datefmt=None):
            dt = datetime.fromtimestamp(record.created, shanghai_tz)
            if datefmt:
                return dt.strftime(datefmt)
            else:
                return dt.strftime('%Y-%m-%d %H:%M:%S')
    
    # 创建格式化器
    formatter = ShanghaiFormatter(
        fmt='%(asctime)s | %(levelname)-8s | %(filename)s:%(lineno)d | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # 配置根日志记录器
    root_logger = logging.getLogger()
    root_logger.setLevel(log_level)
    
    # 清除现有的处理器
    root_logger.handlers.clear()
    
    # 添加控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(log_level)
    console_handler.setFormatter(formatter)
    root_logger.addHandler(console_handler)
    
    # 添加文件处理器
    file_handler = logging.FileHandler(log_filepath, encoding='utf-8')
    file_handler.setLevel(log_level)
    file_handler.setFormatter(formatter)
    root_logger.addHandler(file_handler)
    
    logger = logging.getLogger('KronosPipeline')
    logger.info(f"日志文件: {log_filepath} (上海时区)")
    
    return logger

def setup_ddp():
    """
    初始化分布式数据并行环境
    
    Returns:
        tuple: (rank, world_size, local_rank)
    """
    if not dist.is_available():
        raise RuntimeError("torch.distributed is not available.")

    dist.init_process_group(backend="nccl")
    rank = int(os.environ["RANK"])
    world_size = int(os.environ["WORLD_SIZE"])
    local_rank = int(os.environ["LOCAL_RANK"])
    torch.cuda.set_device(local_rank)
    logger.info(
        f"[DDP Setup] Global Rank: {rank}/{world_size}, "
        f"Local Rank (GPU): {local_rank} on device {torch.cuda.current_device()}"
    )
    return rank, world_size, local_rank

def cleanup_ddp():
    """清理分布式进程组"""
    if dist.is_initialized():
        dist.destroy_process_group()

def set_seed(seed: int, rank: int = 0):
    """
    设置随机种子以确保可重复性
    
    Args:
        seed: 基础种子值
        rank: 进程排名，用于确保不同进程有不同的种子
    """
    actual_seed = seed + rank
    random.seed(actual_seed)
    np.random.seed(actual_seed)
    torch.manual_seed(actual_seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(actual_seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False

def get_model_size(model: torch.nn.Module) -> str:
    """
    计算PyTorch模型中可训练参数的数量
    
    Args:
        model: PyTorch模型
        
    Returns:
        str: 表示模型大小的字符串（例如"175.0B"，"7.1M"，"50.5K"）
    """
    total_params = sum(p.numel() for p in model.parameters() if p.requires_grad)

    if total_params >= 1e9:
        return f"{total_params / 1e9:.1f}B"  # Billions
    elif total_params >= 1e6:
        return f"{total_params / 1e6:.1f}M"  # Millions
    else:
        return f"{total_params / 1e3:.1f}K"  # Thousands

def reduce_tensor(tensor: torch.Tensor, world_size: int, op=dist.ReduceOp.SUM) -> torch.Tensor:
    """
    在分布式设置中减少张量的值
    
    Args:
        tensor: 要减少的张量
        world_size: 总进程数
        op: 归约操作（SUM，AVG等）
        
    Returns:
        torch.Tensor: 归约后的张量
    """
    rt = tensor.clone()
    dist.all_reduce(rt, op=op)
    if op == dist.ReduceOp.SUM and hasattr(dist.ReduceOp, 'AVG'):
        rt /= world_size
    return rt

def format_time(seconds: float) -> str:
    """
    将秒数格式化为人类可读的H:M:S字符串
    
    Args:
        seconds: 总秒数
        
    Returns:
        str: 格式化的时间字符串（例如"0:15:32"）
    """
    return str(timedelta(seconds=int(seconds)))

def create_dataloaders_ddp(config: dict, rank: int, world_size: int):
    """
    为训练和验证创建分布式数据加载器
    
    Args:
        config: 配置参数字典
        rank: 当前进程的全局排名
        world_size: 总进程数
        
    Returns:
        tuple: (train_loader, val_loader, train_dataset, valid_dataset)
    """
    logger.info(f"[Rank {rank}] 创建分布式数据加载器...")
    train_dataset = FinancialDataset('train', config=config)
    valid_dataset = FinancialDataset('val', config=config)
    logger.info(f"[Rank {rank}] 训练数据集大小: {len(train_dataset)}, 验证数据集大小: {len(valid_dataset)}")

    train_sampler = DistributedSampler(train_dataset, num_replicas=world_size, rank=rank, shuffle=True)
    val_sampler = DistributedSampler(valid_dataset, num_replicas=world_size, rank=rank, shuffle=False)

    train_loader = DataLoader(
        train_dataset,
        batch_size=config['batch_size'],
        sampler=train_sampler,
        shuffle=False,  # 由采样器处理洗牌
        num_workers=config.get('num_workers', 2),
        pin_memory=True,
        drop_last=True
    )
    val_loader = DataLoader(
        valid_dataset,
        batch_size=config['batch_size'],
        sampler=val_sampler,
        shuffle=False,
        num_workers=config.get('num_workers', 2),
        pin_memory=True,
        drop_last=False
    )
    logger.info(f"[Rank {rank}] 数据加载器创建完成。每轮训练步数: {len(train_loader)}, 验证步数: {len(val_loader)}")
    return train_loader, val_loader, train_dataset, valid_dataset

def create_dataloaders_cpu(config: dict):
    """
    为CPU训练创建标准（非分布式）数据加载器
    
    Args:
        config: 配置参数字典
        
    Returns:
        tuple: (train_loader, val_loader, train_dataset, valid_dataset)
    """
    logger.info("[CPU] 创建数据加载器...")
    train_dataset = FinancialDataset('train', config=config)
    valid_dataset = FinancialDataset('val', config=config)
    logger.info(f"[CPU] 训练数据集大小: {len(train_dataset)}, 验证数据集大小: {len(valid_dataset)}")

    num_workers = config.get('num_workers', 0) or 0
    pin_memory = False  # 在CPU上，pin_memory没有好处

    train_loader = DataLoader(
        train_dataset,
        batch_size=config['batch_size'],
        shuffle=True,
        num_workers=num_workers,
        pin_memory=pin_memory,
        drop_last=True,
    )

    val_loader = DataLoader(
        valid_dataset,
        batch_size=config['batch_size'],
        shuffle=False,
        num_workers=num_workers,
        pin_memory=pin_memory,
        drop_last=False,
    )

    logger.info(f"[CPU] 数据加载器创建完成。每轮训练步数: {len(train_loader)}, 验证步数: {len(val_loader)}")
    return train_loader, val_loader, train_dataset, valid_dataset

def setup_comet_logger(config):
    """
    设置Comet.ml日志记录器
    
    Args:
        config: 配置参数字典
        
    Returns:
        comet_ml.Experiment或None: Comet日志记录器实例
    """
    if not config.get('use_comet', False):
        return None
        
    try:
        # comet_ml 已在文件顶部导入
        logger.info("初始化Comet.ml日志记录器...")
        comet_logger = comet_ml.Experiment(
            api_key=config['comet_config']['api_key'],
            project_name=config['comet_config']['project_name'],
            workspace=config['comet_config']['workspace'],
        )
        comet_logger.add_tag(config['comet_tag'])
        comet_logger.set_name(config['comet_name'])
        comet_logger.log_parameters(config)
        logger.info("Comet.ml日志记录器初始化成功")
        return comet_logger
    except ImportError:
        logger.warning("无法导入comet_ml，将不使用Comet日志记录")
        return None
    except Exception as e:
        logger.error(f"设置Comet日志记录器时出错: {str(e)}")
        return None

def save_model_checkpoint(model, save_path, is_ddp=False):
    """
    保存模型检查点
    
    Args:
        model: 要保存的模型
        save_path: 保存路径
        is_ddp: 是否是DDP包装的模型
    """
    try:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        if is_ddp:
            model.module.save_pretrained(save_path)
        else:
            model.save_pretrained(save_path)
        logger.info(f"模型已保存到 {save_path}")
        return True
    except Exception as e:
        logger.error(f"保存模型时出错: {str(e)}")
        return False

def save_training_summary(save_dir, summary_data):
    """
    保存训练摘要到JSON文件
    
    Args:
        save_dir: 保存目录
        summary_data: 要保存的摘要数据字典
    """
    try:
        os.makedirs(save_dir, exist_ok=True)
        with open(os.path.join(save_dir, 'training_summary.json'), 'w') as f:
            json.dump(summary_data, f, indent=4)
        logger.info(f"训练摘要已保存到 {os.path.join(save_dir, 'training_summary.json')}")
        return True
    except Exception as e:
        logger.error(f"保存训练摘要时出错: {str(e)}")
        return False

def predict_future_trends(tokenizer, model, test_data, config, device, save_dir=None):
    """
    使用训练好的模型预测未来股票走势
    
    Args:
        tokenizer: 分词模型
        model: 预测模型
        test_data: 测试数据
        config: 配置参数
        device: 设备
        save_dir: 保存目录，如果为None则不保存结果
        
    Returns:
        dict: 预测结果DataFrames的字典
    """
    logger.info("Using predict_future_trends from training_pipeline_utils.py")
    try:
        # 所有需要的模块已在文件顶部导入
        
        # 创建预测数据集
        class PredictionDataset(torch.utils.data.Dataset):
            def __init__(self, data, config):
                self.data = data
                self.config = config
                self.symbols = list(self.data.keys())
                self.feature_list = config.feature_list
                self.time_feature_list = config.time_feature_list
                self.indices = []
                
                # 为每个股票准备最新的数据窗口
                for symbol in self.symbols:
                    df = self.data[symbol].reset_index()
                    # 只使用最新的lookback_window天数据
                    if len(df) >= config.lookback_window:
                        # 生成时间特征
                        df['minute'] = df['datetime'].dt.minute
                        df['hour'] = df['datetime'].dt.hour
                        df['weekday'] = df['datetime'].dt.weekday
                        df['day'] = df['datetime'].dt.day
                        df['month'] = df['datetime'].dt.month
                        self.data[symbol] = df
                        
                        # 使用最新的数据窗口
                        start_idx = len(df) - config.lookback_window
                        timestamp = df.iloc[-1]['datetime']
                        self.indices.append((symbol, start_idx, timestamp))
            
            def __len__(self):
                return len(self.indices)
            
            def __getitem__(self, idx):
                symbol, start_idx, timestamp = self.indices[idx]
                df = self.data[symbol]
                
                # 获取上下文窗口
                context_end = start_idx + self.config.lookback_window
                context_df = df.iloc[start_idx:context_end]
                
                # 生成未来10个工作日的时间戳特征
                last_date = context_df.iloc[-1]['datetime']
                future_dates = get_future_business_days(last_date, self.config.predict_window)
                future_features = pd.DataFrame({
                    'datetime': future_dates,
                    'minute': [d.minute for d in future_dates],
                    'hour': [d.hour for d in future_dates],
                    'weekday': [d.weekday() for d in future_dates],
                    'day': [d.day for d in future_dates],
                    'month': [d.month for d in future_dates]
                })
                
                # 提取特征
                x = context_df[self.feature_list].values.astype(np.float32)
                x_stamp = context_df[self.time_feature_list].values.astype(np.float32)
                y_stamp = future_features[self.time_feature_list].values.astype(np.float32)
                
                # 标准化
                x_mean, x_std = np.mean(x, axis=0), np.std(x, axis=0)
                x = (x - x_mean) / (x_std + 1e-5)
                x = np.clip(x, -self.config.clip, self.config.clip)
                
                return torch.from_numpy(x), torch.from_numpy(x_stamp), torch.from_numpy(y_stamp), symbol, timestamp
        
        # 创建预测数据集和数据加载器
        pred_dataset = PredictionDataset(test_data, config)
        
        # 定义collate函数
        def collate_fn(batch):
            x, x_stamp, y_stamp, symbols, timestamps = zip(*batch)
            x_batch = torch.stack(x, dim=0)
            x_stamp_batch = torch.stack(x_stamp, dim=0)
            y_stamp_batch = torch.stack(y_stamp, dim=0)
            return x_batch, x_stamp_batch, y_stamp_batch, list(symbols), list(timestamps)
        
        pred_loader = torch.utils.data.DataLoader(
            pred_dataset,
            batch_size=max(1, config.batch_size // config.inference_sample_count),
            shuffle=False,
            num_workers=0,
            collate_fn=collate_fn
        )
        
        # 执行预测
        results = defaultdict(list)
        detailed_results = []  # 存储详细的预测结果
        
        with torch.no_grad():
            for x, x_stamp, y_stamp, symbols, timestamps in pred_loader:
                # 使用自回归推理进行预测
                preds = auto_regressive_inference(
                    tokenizer, model, x.to(device), x_stamp.to(device), y_stamp.to(device),
                    max_context=config.max_context, 
                    pred_len=config.predict_window, 
                    clip=config.clip,
                    T=config.inference_T, 
                    top_k=config.inference_top_k, 
                    top_p=config.inference_top_p, 
                    sample_count=config.inference_sample_count
                )
                
                # 只保留预测窗口的数据
                preds = preds[:, -config.predict_window:, :]
                
                # 获取特征在 feature_list 中的索引
                # feature_list = ['open', 'high', 'low', 'close', 'vol', 'amt']
                feature_names = ['open', 'high', 'low', 'close', 'volume', 'amount']
                feature_indices = {
                    'open': 0,
                    'high': 1,
                    'low': 2,
                    'close': 3,
                    'volume': 4,
                    'amount': 5
                }
                
                # 获取最后一天的所有特征值
                last_day_features = {}
                for feat_name, feat_idx in feature_indices.items():
                    last_day_features[feat_name] = x[:, -1, feat_idx].numpy()
                
                # 计算基于close价格的信号类型（用于排序和筛选）
                last_day_close = last_day_features['close']
                signals = {
                    'last': preds[:, -1, 3] - last_day_close,
                    'mean': np.mean(preds[:, :, 3], axis=1) - last_day_close,
                    'max': np.max(preds[:, :, 3], axis=1) - last_day_close,
                    'min': np.min(preds[:, :, 3], axis=1) - last_day_close,
                }
                
                # 收集结果
                for i in range(len(symbols)):
                    for sig_type, sig_values in signals.items():
                        results[sig_type].append((timestamps[i], symbols[i], sig_values[i]))
                    
                    # 收集详细预测结果（每支股票未来N天的所有特征预测）
                    symbol = symbols[i]
                    
                    # 为每支股票创建一个详细记录
                    detail_record = {
                        'stock_code': symbol,
                        'current_open': last_day_features['open'][i],
                        'current_high': last_day_features['high'][i],
                        'current_low': last_day_features['low'][i],
                        'current_close': last_day_features['close'][i],
                        'current_volume': last_day_features['volume'][i],
                        'current_amount': last_day_features['amount'][i],
                    }
                    
                    # 获取未来日期
                    last_date = timestamps[i]
                    future_dates = get_future_business_days(last_date, config.predict_window)
                    
                    # 添加未来每一天的所有特征预测
                    for day_idx in range(config.predict_window):
                        day_num = day_idx + 1
                        
                        # 保存预测日期
                        pred_date = future_dates[day_idx].strftime('%Y-%m-%d') if day_idx < len(future_dates) else None
                        detail_record[f'day_{day_num}_date'] = pred_date
                        
                        # 预测的所有特征值
                        for feat_name, feat_idx in feature_indices.items():
                            pred_value = preds[i, day_idx, feat_idx]
                            current_value = last_day_features[feat_name][i]
                            
                            # 保存预测值
                            detail_record[f'day_{day_num}_{feat_name}'] = pred_value
                            
                            # 计算变化和变化百分比（对于价格和成交量）
                            if feat_name in ['open', 'high', 'low', 'close']:
                                change = pred_value - current_value
                                change_pct = (change / current_value) * 100 if current_value != 0 else 0
                                detail_record[f'day_{day_num}_{feat_name}_change'] = change
                                detail_record[f'day_{day_num}_{feat_name}_change_pct'] = change_pct
                            elif feat_name in ['volume', 'amount']:
                                change = pred_value - current_value
                                change_pct = (change / current_value) * 100 if current_value != 0 else 0
                                detail_record[f'day_{day_num}_{feat_name}_change'] = change
                                detail_record[f'day_{day_num}_{feat_name}_change_pct'] = change_pct
                    
                    detailed_results.append(detail_record)
        
        # 处理预测结果
        logger.info("处理预测结果...")
        prediction_dfs = {}
        for sig_type, records in results.items():
            df = pd.DataFrame(records, columns=['datetime', 'instrument', 'score'])
            pivot_df = df.pivot_table(index='datetime', columns='instrument', values='score')
            prediction_dfs[sig_type] = pivot_df.sort_index()
        
        # 创建详细的预测DataFrame
        detailed_df = pd.DataFrame(detailed_results)
        prediction_dfs['detailed'] = detailed_df
        
        # 保存预测结果
        if save_dir is not None:
            prediction_dir = os.path.join(save_dir, "predictions")
            os.makedirs(prediction_dir, exist_ok=True)
            timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 保存pkl文件（包含所有结果）
            prediction_file = os.path.join(prediction_dir, f"prediction_{timestamp_str}.pkl")
            with open(prediction_file, 'wb') as f:
                pickle.dump(prediction_dfs, f)
            logger.info(f"pkl结果已保存到: {prediction_file}")
            
            # 保存Excel文件（主要是详细预测结果）
            excel_file = os.path.join(prediction_dir, f"prediction_{timestamp_str}.xlsx")
            
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    # 1. 保存详细预测结果（包含每支股票未来N天的所有特征数据）
                    if 'detailed' in prediction_dfs and len(detailed_df) > 0:
                        # 重命名列以便更易理解
                        column_mapping = {
                            'stock_code': '股票代码',
                            'current_open': '当前开盘价',
                            'current_high': '当前最高价',
                            'current_low': '当前最低价',
                            'current_close': '当前收盘价',
                            'current_volume': '当前成交量',
                            'current_amount': '当前成交额',
                        }
                        
                        # 特征名称映射
                        feature_name_map = {
                            'open': '开盘价',
                            'high': '最高价',
                            'low': '最低价',
                            'close': '收盘价',
                            'volume': '成交量',
                            'amount': '成交额'
                        }
                        
                        # 为每一天的每个特征添加列名映射
                        for day_idx in range(config.predict_window):
                            day_num = day_idx + 1
                            # 获取对应的预测日期
                            pred_date = detailed_df[f'day_{day_num}_date'].iloc[0] if f'day_{day_num}_date' in detailed_df.columns else None
                            date_str = f'({pred_date})' if pred_date else ''
                            
                            # 添加日期列映射
                            if f'day_{day_num}_date' in detailed_df.columns:
                                column_mapping[f'day_{day_num}_date'] = f'第{day_num}天日期'
                            
                            for feat_key, feat_name in feature_name_map.items():
                                column_mapping[f'day_{day_num}_{feat_key}'] = f'第{day_num}天{date_str}预测{feat_name}'
                                column_mapping[f'day_{day_num}_{feat_key}_change'] = f'第{day_num}天{date_str}{feat_name}变化'
                                column_mapping[f'day_{day_num}_{feat_key}_change_pct'] = f'第{day_num}天{date_str}{feat_name}涨跌幅(%)'
                        
                        # 创建显示用的DataFrame
                        display_df = detailed_df.copy()
                        display_df = display_df.rename(columns=column_mapping)
                        
                        # 保存到Excel
                        display_df.to_excel(writer, sheet_name='详细预测', index=False, startrow=1)
                        
                        # 获取工作表对象
                        worksheet = writer.sheets['详细预测']
                        
                        # 添加标题
                        worksheet['A1'] = f'股票预测结果 - 生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                        worksheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
                        worksheet['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        worksheet.merge_cells(f'A1:{get_column_letter(len(display_df.columns))}1')
                        
                        # 设置表头样式
                        header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
                        header_font = Font(bold=True, size=11)
                        border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for col_idx, col_name in enumerate(display_df.columns, start=1):
                            cell = worksheet.cell(row=2, column=col_idx)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = border
                        
                        # 设置数据行样式和格式
                        for row_idx in range(3, len(display_df) + 3):
                            for col_idx in range(1, len(display_df.columns) + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.border = border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                
                                # 为涨跌幅列添加颜色
                                col_name = display_df.columns[col_idx - 1]
                                if '涨跌幅' in col_name:
                                    value = cell.value
                                    if value is not None and isinstance(value, (int, float)):
                                        if value > 0:
                                            cell.font = Font(color='FF0000', bold=True)  # 红色表示上涨
                                        elif value < 0:
                                            cell.font = Font(color='00B050', bold=True)  # 绿色表示下跌
                                        # 格式化为两位小数
                                        cell.number_format = '0.00'
                                elif '价格' in col_name or '收盘价' in col_name:
                                    # 价格列格式化为两位小数
                                    if cell.value is not None and isinstance(cell.value, (int, float)):
                                        cell.number_format = '0.00'
                        
                        # 自动调整列宽
                        for col_idx, col_name in enumerate(display_df.columns, start=1):
                            max_length = len(str(col_name)) + 2
                            for row_idx in range(3, min(len(display_df) + 3, 100)):  # 只检查前100行
                                try:
                                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                                    if cell_value is not None:
                                        max_length = max(max_length, len(str(cell_value)) + 2)
                                except:
                                    pass
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length, 20)
                        
                        logger.info(f"详细预测表已保存，包含 {len(detailed_df)} 支股票")
                    
                    # 2. 创建各种排行榜
                    if 'detailed' in prediction_dfs and len(detailed_df) > 0:
                        # 收盘价涨跌幅排行榜（第N天）
                        if f'day_{config.predict_window}_close_change_pct' in detailed_df.columns:
                            cols_to_select = ['stock_code', 'current_close', 
                                          f'day_{config.predict_window}_close', 
                                          f'day_{config.predict_window}_close_change_pct']
                            
                            # 如果有日期列，也加入选择
                            if f'day_{config.predict_window}_date' in detailed_df.columns:
                                cols_to_select.append(f'day_{config.predict_window}_date')
                                
                            ranking_df = detailed_df[cols_to_select].copy()
                            ranking_df = ranking_df.sort_values(f'day_{config.predict_window}_close_change_pct', ascending=False)
                        
                        # 获取预测日期
                        pred_date = ranking_df[f'day_{config.predict_window}_date'].iloc[0] if f'day_{config.predict_window}_date' in ranking_df.columns else None
                        date_str = f'({pred_date})' if pred_date else ''
                        
                        # 重命名列
                        col_names = ['股票代码', '当前收盘价', f'第{config.predict_window}天{date_str}预测价格', f'第{config.predict_window}天{date_str}涨跌幅(%)']
                        if f'day_{config.predict_window}_date' in ranking_df.columns:
                            ranking_df = ranking_df.drop(columns=[f'day_{config.predict_window}_date'])
                            
                        ranking_df.columns = col_names
                        
                        # 添加排名
                        ranking_df.insert(0, '排名', range(1, len(ranking_df) + 1))
                        
                        # 保存到Excel
                        ranking_df.to_excel(writer, sheet_name='涨跌幅排行榜', index=False, startrow=1)
                        
                        # 美化排行榜
                        worksheet = writer.sheets['涨跌幅排行榜']
                        # 获取预测日期
                        pred_date = detailed_df[f'day_{config.predict_window}_date'].iloc[0] if f'day_{config.predict_window}_date' in detailed_df.columns else None
                        date_str = f' {pred_date} ' if pred_date else ' '
                        worksheet['A1'] = f'股票涨跌幅排行榜 (第{config.predict_window}天{date_str}预测)'
                        worksheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
                        worksheet['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                        worksheet.merge_cells(f'A1:{get_column_letter(len(ranking_df.columns))}1')
                        
                        # 设置表头样式
                        for col_idx in range(1, len(ranking_df.columns) + 1):
                            cell = worksheet.cell(row=2, column=col_idx)
                            cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = border
                        
                        # 设置数据行样式
                        for row_idx in range(3, len(ranking_df) + 3):
                            for col_idx in range(1, len(ranking_df.columns) + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.border = border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                
                                # 为涨跌幅列添加颜色
                                if col_idx == len(ranking_df.columns):  # 最后一列是涨跌幅
                                    value = cell.value
                                    if value is not None and isinstance(value, (int, float)):
                                        if value > 0:
                                            cell.font = Font(color='FF0000', bold=True)
                                            cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                                        elif value < 0:
                                            cell.font = Font(color='00B050', bold=True)
                                            cell.fill = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')
                                        cell.number_format = '0.00'
                                elif col_idx > 1:  # 价格列
                                    if cell.value is not None and isinstance(cell.value, (int, float)):
                                        cell.number_format = '0.00'
                        
                        # 自动调整列宽
                        for col_idx in range(1, len(ranking_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = 18
                        
                        logger.info(f"收盘价涨跌幅排行榜已保存")
                        
                        # 为每个特征创建简化视图（只显示预测值和涨跌幅）
                        feature_views = {
                            'close': '收盘价',
                            'open': '开盘价',
                            'high': '最高价',
                            'low': '最低价',
                            'volume': '成交量',
                            'amount': '成交额'
                        }
                        
                        for feat_key, feat_name in feature_views.items():
                            # 选择相关列
                            cols_to_include = ['stock_code', f'current_{feat_key}']
                            for day_idx in range(config.predict_window):
                                day_num = day_idx + 1
                                cols_to_include.append(f'day_{day_num}_{feat_key}')
                                cols_to_include.append(f'day_{day_num}_{feat_key}_change_pct')
                            
                            # 检查列是否存在
                            existing_cols = [col for col in cols_to_include if col in detailed_df.columns]
                            if len(existing_cols) > 2:  # 至少有股票代码和当前值
                                feat_df = detailed_df[existing_cols].copy()
                                
                                # 重命名列
                                col_rename = {
                                    'stock_code': '股票代码',
                                    f'current_{feat_key}': f'当前{feat_name}'
                                }
                                for day_idx in range(config.predict_window):
                                    day_num = day_idx + 1
                                    # 获取对应的预测日期
                                    pred_date = detailed_df[f'day_{day_num}_date'].iloc[0] if f'day_{day_num}_date' in detailed_df.columns else None
                                    date_str = f'({pred_date})' if pred_date else ''
                                    
                                    col_rename[f'day_{day_num}_{feat_key}'] = f'第{day_num}天{date_str}{feat_name}'
                                    col_rename[f'day_{day_num}_{feat_key}_change_pct'] = f'第{day_num}天{date_str}涨跌幅(%)'
                                    # 如果有日期列也添加到列名映射中
                                    if f'day_{day_num}_date' in detailed_df.columns and f'day_{day_num}_date' in feat_df.columns:
                                        col_rename[f'day_{day_num}_date'] = f'第{day_num}天日期'
                                
                                feat_df = feat_df.rename(columns=col_rename)
                                
                                # 保存到Excel
                                sheet_name = f'{feat_name}预测'
                                feat_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
                                
                                # 美化工作表
                                worksheet = writer.sheets[sheet_name]
                                worksheet['A1'] = f'{feat_name}预测详情'
                                worksheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
                                worksheet['A1'].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
                                worksheet.merge_cells(f'A1:{get_column_letter(len(feat_df.columns))}1')
                                
                                # 设置表头样式
                                for col_idx in range(1, len(feat_df.columns) + 1):
                                    cell = worksheet.cell(row=2, column=col_idx)
                                    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                                    cell.font = Font(bold=True)
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    cell.border = border
                                
                                # 设置数据格式
                                for row_idx in range(3, len(feat_df) + 3):
                                    for col_idx in range(1, len(feat_df.columns) + 1):
                                        cell = worksheet.cell(row=row_idx, column=col_idx)
                                        cell.border = border
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        
                                        col_name = feat_df.columns[col_idx - 1]
                                        if '涨跌幅' in col_name:
                                            value = cell.value
                                            if value is not None and isinstance(value, (int, float)):
                                                if value > 0:
                                                    cell.font = Font(color='FF0000', bold=True)
                                                elif value < 0:
                                                    cell.font = Font(color='00B050', bold=True)
                                                cell.number_format = '0.00'
                                        elif col_idx > 1:  # 数值列
                                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                                if feat_key in ['volume', 'amount']:
                                                    cell.number_format = '#,##0'  # 千分位
                                                else:
                                                    cell.number_format = '0.00'
                                
                                # 自动调整列宽
                                for col_idx in range(1, len(feat_df.columns) + 1):
                                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
                                
                                logger.info(f"{feat_name}预测表已保存")
                    
                    # 3. 保存各种信号类型的结果
                    for sig_type, df in prediction_dfs.items():
                        if sig_type != 'detailed':
                            df.to_excel(writer, sheet_name=f'信号_{sig_type}')
                
                logger.info(f"✓ Excel结果已保存到: {excel_file}")
                
            except ImportError:
                logger.warning("未安装openpyxl库，无法保存Excel文件")
                logger.warning("请运行: pip install openpyxl")
                logger.info("您仍然可以使用pkl文件查看预测结果")
            except Exception as e:
                logger.warning(f"保存Excel文件时出错: {str(e)}")
                logger.info("但pkl文件已成功保存，您可以使用pkl文件查看预测结果")
        
        return prediction_dfs
    except Exception as e:
        logger.error(f"预测时出错: {str(e)}")
        # traceback 已在文件顶部导入
        logger.error(traceback.format_exc())
        return None

def save_pipeline_config(config, save_dir):
    """
    保存完整的流水线配置
    
    Args:
        config: 配置对象
        save_dir: 保存目录
    """
    try:
        os.makedirs(save_dir, exist_ok=True)
        config_dict = {k: v for k, v in config.__dict__.items() if not k.startswith('_')}
        
        # 处理不可序列化的对象
        for key in list(config_dict.keys()):
            try:
                json.dumps({key: config_dict[key]})
            except (TypeError, OverflowError):
                config_dict[key] = str(config_dict[key])
        
        with open(os.path.join(save_dir, 'pipeline_config.json'), 'w') as f:
            json.dump(config_dict, f, indent=4)
        logger.info(f"流水线配置已保存到 {os.path.join(save_dir, 'pipeline_config.json')}")
        return True
    except Exception as e:
        logger.error(f"保存流水线配置时出错: {str(e)}")
        return False


def evaluate_model_on_test_data(model, tokenizer, test_data, config, device):
    """
    在测试数据上评估模型
    
    Args:
        model: 预测模型
        tokenizer: 分词模型
        test_data: 测试数据
        config: 配置参数
        device: 设备
        
    Returns:
        float: 平均验证损失
    """
    logger.info("Using evaluate_model_on_test_data from training_pipeline_utils.py")
    tot_val_loss = 0.0
    val_batches_processed = 0
    
    # 使用测试数据创建一个临时数据集
    # 使用与FinancialDataset完全相同的数据处理逻辑，确保格式一致性
    class TestDataset(FinancialDataset):
        def __init__(self, test_data, config):
            self.config = config
            # 关键修复：不修改传入的test_data，而是直接从中读取后保存到self.data
            # 这样每次创建TestDataset时都从原始test_data开始处理
            self.data = {}
            self.window = self.config.lookback_window + self.config.predict_window + 1
            self.symbols = list(test_data.keys())
            self.feature_list = self.config.feature_list
            self.time_feature_list = self.config.time_feature_list
            self.indices = []
            self.py_rng = random.Random(self.config.seed)
            
            # 预处理数据 - 与FinancialDataset保持完全一致的逻辑
            for symbol in self.symbols:
                # reset_index()将datetime索引转换为列
                # 从原始test_data中获取数据，而不是self.data
                df = test_data[symbol].reset_index()
                series_len = len(df)
                num_samples = series_len - self.window + 1
                
                if num_samples > 0:
                    # 生成时间特征 - 与FinancialDataset相同
                    df['minute'] = df['datetime'].dt.minute
                    df['hour'] = df['datetime'].dt.hour
                    df['weekday'] = df['datetime'].dt.weekday
                    df['day'] = df['datetime'].dt.day
                    df['month'] = df['datetime'].dt.month
                    
                    # 只保留需要的特征列 - 与FinancialDataset相同
                    # 现在保存到self.data中
                    self.data[symbol] = df[self.feature_list + self.time_feature_list]
                    
                    # 添加所有有效的起始索引
                    for i in range(num_samples):
                        self.indices.append((symbol, i))
            
            # 限制测试样本数量
            self.n_samples = min(self.config.n_val_iter, len(self.indices))
    
    # 创建测试数据集和数据加载器
    test_dataset = TestDataset(test_data, config)
    test_loader = DataLoader(
        test_dataset,
        batch_size=config.batch_size,
        shuffle=False,
        num_workers=0,
        pin_memory=False
    )
    
    # 在测试集上评估模型
    model.eval()
    with torch.no_grad():
        for batch_x, batch_x_stamp in test_loader:
            batch_x = batch_x.squeeze(0).to(device)
            batch_x_stamp = batch_x_stamp.squeeze(0).to(device)
            
            token_seq_0, token_seq_1 = tokenizer.encode(batch_x, half=True)
            token_in = [token_seq_0[:, :-1], token_seq_1[:, :-1]]
            token_out = [token_seq_0[:, 1:], token_seq_1[:, 1:]]
            
            logits = model(token_in[0], token_in[1], batch_x_stamp[:, :-1, :])
            val_loss, _, _ = model.head.compute_loss(logits[0], logits[1], token_out[0], token_out[1])
            
            tot_val_loss += val_loss.item()
            val_batches_processed += 1
    
    avg_val_loss = tot_val_loss / val_batches_processed if val_batches_processed > 0 else float('inf')
    return avg_val_loss


def evaluate_tokenizer_on_test_data(tokenizer_path, test_data, config, device, is_remote=False):
    """在测试数据上评估分词模型
    
    Args:
        tokenizer_path: 分词模型路径或远程模型名称
        test_data: 测试数据
        config: 配置参数
        device: 设备
        is_remote: 是否从远程加载模型，默认为False
        
    Returns:
        float: 测试损失
    """
    try:
        # 加载分词模型，根据is_remote参数决定从本地或远程加载
        if is_remote:
            logger.info(f"从远程加载分词模型: {tokenizer_path}")
            tokenizer = KronosTokenizer.from_pretrained(tokenizer_path)
        else:
            logger.info(f"从本地加载分词模型: {tokenizer_path}")
            tokenizer = KronosTokenizer.from_pretrained(tokenizer_path, local_files_only=True)
        tokenizer.eval().to(device)
        
        # 创建测试数据集
        class TokenizerTestDataset(torch.utils.data.Dataset):
            def __init__(self, data, config):
                self.data = data
                self.config = config
                self.symbols = list(self.data.keys())
                self.feature_list = config.feature_list
                self.indices = []
                
                # 为每个股票准备测试数据窗口
                for symbol in self.symbols:
                    df = self.data[symbol].reset_index()
                    if len(df) >= config.lookback_window:
                        for i in range(0, len(df) - config.lookback_window + 1, config.lookback_window // 2):
                            self.indices.append((symbol, i))
            
            def __len__(self):
                return min(self.config.n_val_iter, len(self.indices))
            
            def __getitem__(self, idx):
                symbol, start_idx = self.indices[idx % len(self.indices)]
                df = self.data[symbol]
                
                # 获取数据窗口
                end_idx = min(start_idx + self.config.lookback_window, len(df))
                x = df.iloc[start_idx:end_idx][self.feature_list].values.astype(np.float32)
                
                # 标准化
                x_mean, x_std = np.mean(x, axis=0), np.std(x, axis=0)
                x = (x - x_mean) / (x_std + 1e-5)
                x = np.clip(x, -self.config.clip, self.config.clip)
                
                return torch.from_numpy(x)
        
        # 创建测试数据集和数据加载器
        test_dataset = TokenizerTestDataset(test_data, config)
        test_loader = torch.utils.data.DataLoader(
            test_dataset,
            batch_size=config.batch_size,
            shuffle=False,
            num_workers=0
        )
        
        # 评估模型
        tot_test_loss = 0.0
        test_batches = 0
        
        with torch.no_grad():
            for batch_x in test_loader:
                batch_x = batch_x.to(device)
                zs, _, _, _ = tokenizer(batch_x)
                _, z = zs
                test_loss = torch.nn.functional.mse_loss(z, batch_x)
                tot_test_loss += test_loss.item()
                test_batches += 1
        
        avg_test_loss = tot_test_loss / test_batches if test_batches > 0 else float('inf')
        return avg_test_loss
    
    except Exception as e:
        logger.error(f"评估分词模型时出错: {str(e)}")
        # traceback 已在文件顶部导入
        logger.error(traceback.format_exc())
        return float('inf')


def evaluate_predictor_on_test_data(predictor_path, tokenizer_path, test_data, config, device, is_remote_predictor=False, is_remote_tokenizer=False):
    """在测试数据上评估预测模型
    
    Args:
        predictor_path: 预测模型路径或远程模型名称
        tokenizer_path: 分词模型路径或远程模型名称
        test_data: 测试数据
        config: 配置参数
        device: 设备
        is_remote_predictor: 预测模型是否从远程加载，默认为False
        is_remote_tokenizer: 分词模型是否从远程加载，默认为False
        
    Returns:
        float: 测试损失
    """
    try:
        # 加载分词模型，根据is_remote_tokenizer参数决定从本地或远程加载
        if is_remote_tokenizer:
            logger.info(f"从远程加载分词模型: {tokenizer_path}")
            tokenizer = KronosTokenizer.from_pretrained(tokenizer_path)
        else:
            logger.info(f"从本地加载分词模型: {tokenizer_path}")
            tokenizer = KronosTokenizer.from_pretrained(tokenizer_path, local_files_only=True)
        tokenizer.eval().to(device)
        
        # 加载预测模型，根据is_remote_predictor参数决定从本地或远程加载
        if is_remote_predictor:
            logger.info(f"从远程加载预测模型: {predictor_path}")
            model = Kronos.from_pretrained(predictor_path)
        else:
            logger.info(f"从本地加载预测模型: {predictor_path}")
            model = Kronos.from_pretrained(predictor_path, local_files_only=True)
        model.eval().to(device)
        
        # 使用工具函数评估模型
        test_loss = evaluate_model_on_test_data(model, tokenizer, test_data, config, device)
        return test_loss
    except Exception as e:
        logger.error(f"评估预测模型时出错: {str(e)}")
        # traceback 已在文件顶部导入
        logger.error(traceback.format_exc())
        return float('inf')


def evaluate_and_select_best_model(model_type, model_paths, test_data, config, device, save_path=None):
    """
    评估多个模型并选择最佳模型的通用框架
    
    Args:
        model_type: 模型类型，'tokenizer' 或 'predictor'
        model_paths: 模型路径列表，格式为 [{'name': 'model_name', 'path': 'model_path', 'tokenizer_path': 'tokenizer_path'}]
                     对于tokenizer类型，tokenizer_path可以省略
        test_data: 测试数据
        config: 配置参数
        device: 设备
        save_path: 可选，保存最佳模型的路径
        
    Returns:
        tuple: (最佳模型名称, 最佳模型路径, 最佳验证损失)
    """
    try:
        logger.info(f"开始评估{model_type}模型...")
        best_model_name = None
        best_model_path = None
        best_val_loss = float('inf')
        
        for model_info in model_paths:
            model_name = model_info['name']
            model_path = model_info['path']
            
            # 检查模型路径是否存在
            if not os.path.exists(model_path):
                logger.warning(f"{model_name} 模型路径不存在，跳过评估: {model_path}")
                continue
            
            # 根据模型类型选择评估方法
            if model_type == 'tokenizer':
                logger.info(f"评估分词模型 {model_name}: {model_path}")
                val_loss = evaluate_tokenizer_on_test_data(model_path, test_data, config, device)
            elif model_type == 'predictor':
                tokenizer_path = model_info.get('tokenizer_path')
                if not tokenizer_path or not os.path.exists(tokenizer_path):
                    logger.warning(f"{model_name} 分词器路径不存在，跳过评估: {tokenizer_path}")
                    continue
                logger.info(f"评估预测模型 {model_name}: {model_path}")
                val_loss = evaluate_predictor_on_test_data(model_path, tokenizer_path, test_data, config, device)
            else:
                logger.error(f"不支持的模型类型: {model_type}")
                return None, None, float('inf')
                
            logger.info(f"{model_name} 模型在测试集上的损失: {val_loss:.4f}")
            
            # 更新最佳模型
            if val_loss < best_val_loss:
                best_val_loss = val_loss
                best_model_name = model_name
                best_model_path = model_path
                
                # 如果提供了保存路径，复制到最终路径
                if save_path:
                    # 复制模型到最终路径
                    logger.info(f"将 {model_name} 模型复制为最佳模型")
                    os.makedirs(save_path, exist_ok=True)
                    os.system(f"cp -r {model_path}/* {save_path}/")
        
        logger.info(f"模型评估完成，最佳模型: {best_model_name}，测试损失: {best_val_loss:.4f}")
        return best_model_name, best_model_path, best_val_loss
    except Exception as e:
        logger.error(f"模型评估过程中出错: {str(e)}")
        # traceback 已在文件顶部导入
        logger.error(traceback.format_exc())
        return None, None, float('inf')


class ModelEvaluationHistory:
    """模型评估历史记录管理类"""
    
    def __init__(self, history_dir=None):
        """
        初始化模型评估历史记录管理器
        
        Args:
            history_dir: 历史记录目录，默认为None
        """
        self.history_dir = history_dir
        if history_dir:
            os.makedirs(history_dir, exist_ok=True)
            self.history_file = os.path.join(history_dir, 'model_history.json')
            self.metrics_file = os.path.join(history_dir, 'model_metrics.csv')
        else:
            self.history_file = None
            self.metrics_file = None
        
        # 初始化历史记录
        self.history = self._load_history()
    
    def _load_history(self):
        """加载历史记录"""
        if not self.history_file or not os.path.exists(self.history_file):
            return {
                'tokenizer': {
                    'daily_records': [],
                    'best_model': None,
                    'best_loss': float('inf')
                },
                'predictor': {
                    'daily_records': [],
                    'best_model': None,
                    'best_loss': float('inf')
                }
            }
        
        try:
            with open(self.history_file, 'r') as f:
                history = json.load(f)
            logger.info(f"加载模型评估历史记录: {self.history_file}")
            return history
        except Exception as e:
            logger.error(f"加载模型评估历史记录时出错: {str(e)}")
            logger.error(traceback.format_exc())
            return {
                'tokenizer': {'daily_records': [], 'best_model': None, 'best_loss': float('inf')},
                'predictor': {'daily_records': [], 'best_model': None, 'best_loss': float('inf')}
            }
    
    def save_history(self):
        """保存历史记录"""
        if not self.history_file:
            return False
        
        try:
            with open(self.history_file, 'w') as f:
                json.dump(self.history, f, indent=4)
            logger.info(f"保存模型评估历史记录: {self.history_file}")
            return True
        except Exception as e:
            logger.error(f"保存模型评估历史记录时出错: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    def add_evaluation_record(self, model_type, model_results, best_model_name):
        """
        添加评估记录
        
        Args:
            model_type: 模型类型，'tokenizer' 或 'predictor'
            model_results: 模型评估结果
            best_model_name: 最佳模型名称
            
        Returns:
            bool: 是否成功添加
        """
        try:
            # 创建当前记录
            current_record = {
                'date': datetime.now().strftime('%Y-%m-%d'),
                'models': model_results,
                'best_model': best_model_name,
                'best_loss': model_results[best_model_name]['loss'] if best_model_name in model_results else None
            }
            
            # 添加到历史记录
            self.history[model_type]['daily_records'].append(current_record)
            
            # 更新全局最佳模型
            if best_model_name and model_results[best_model_name]['loss'] < self.history[model_type]['best_loss']:
                self.history[model_type]['best_model'] = best_model_name
                self.history[model_type]['best_loss'] = model_results[best_model_name]['loss']
                self.history[model_type]['best_date'] = current_record['date']
                self.history[model_type]['best_path'] = model_results[best_model_name]['path']
                logger.info(f"更新{model_type}全局最佳模型: {best_model_name}, 损失: {self.history[model_type]['best_loss']:.4f}")
            
            # 更新指标CSV文件
            self._update_metrics_csv(model_type, model_results, best_model_name, current_record['date'])
            
            # 保存历史记录
            return self.save_history()
        except Exception as e:
            logger.error(f"添加评估记录时出错: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    def _update_metrics_csv(self, model_type, model_results, best_model_name, date):
        """更新指标CSV文件"""
        if not self.metrics_file:
            return
        
        try:
            # 准备数据
            metrics = []
            for model_name, result in model_results.items():
                metrics.append({
                    'date': date,
                    'model_type': model_type,
                    'model_name': model_name,
                    'loss': result['loss'],
                    'is_best': model_name == best_model_name,
                    'path': result['path']
                })
            
            # 加载现有CSV或创建新的
            if os.path.exists(self.metrics_file):
                try:
                    df = pd.read_csv(self.metrics_file)
                except:
                    df = pd.DataFrame(columns=['date', 'model_type', 'model_name', 'loss', 'is_best', 'path'])
            else:
                df = pd.DataFrame(columns=['date', 'model_type', 'model_name', 'loss', 'is_best', 'path'])
            
            # 添加新记录
            df = pd.concat([df, pd.DataFrame(metrics)], ignore_index=True)
            
            # 保存CSV
            df.to_csv(self.metrics_file, index=False)
            logger.info(f"更新模型评估指标CSV: {self.metrics_file}")
        except Exception as e:
            logger.error(f"更新指标CSV时出错: {str(e)}")
            logger.error(traceback.format_exc())
    
    def get_best_model(self, model_type):
        """
        获取最佳模型信息
        
        Args:
            model_type: 模型类型，'tokenizer' 或 'predictor'
            
        Returns:
            dict: 最佳模型信息
        """
        return {
            'model': self.history[model_type]['best_model'],
            'loss': self.history[model_type]['best_loss'],
            'date': self.history[model_type].get('best_date'),
            'path': self.history[model_type].get('best_path')
        }
    
    def get_recent_records(self, model_type, days=7):
        """
        获取最近的记录
        
        Args:
            model_type: 模型类型，'tokenizer' 或 'predictor'
            days: 天数，默认为7
            
        Returns:
            list: 最近的记录列表
        """
        records = self.history[model_type]['daily_records']
        if not records:
            return []
        
        # 按日期排序并返回最近的记录
        records.sort(key=lambda x: x['date'], reverse=True)
        return records[:days]


def load_model_evaluation_history(history_file):
    """
    加载模型评估历史记录（兼容旧版本）
    
    Args:
        history_file: 历史记录文件路径
        
    Returns:
        dict: 历史记录数据
    """
    try:
        if os.path.exists(history_file):
            with open(history_file, 'r') as f:
                history = json.load(f)
            logger.info(f"加载模型评估历史记录: {history_file}")
            return history
        else:
            logger.info(f"历史记录文件不存在，创建新的历史记录: {history_file}")
            return {
                'tokenizer': {
                    'daily_records': [],
                    'best_model': None,
                    'best_loss': float('inf')
                },
                'predictor': {
                    'daily_records': [],
                    'best_model': None,
                    'best_loss': float('inf')
                }
            }
    except Exception as e:
        logger.error(f"加载模型评估历史记录时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return {
            'tokenizer': {'daily_records': [], 'best_model': None, 'best_loss': float('inf')},
            'predictor': {'daily_records': [], 'best_model': None, 'best_loss': float('inf')}
        }


def save_model_evaluation_history(history_file, model_type, model_results, best_model_name):
    """
    保存模型评估历史记录（兼容旧版本）
    
    Args:
        history_file: 历史记录文件路径
        model_type: 模型类型，'tokenizer' 或 'predictor'
        model_results: 模型评估结果
        best_model_name: 最佳模型名称
        
    Returns:
        bool: 是否成功保存
    """
    # 创建历史记录管理器并添加记录
    history_dir = os.path.dirname(history_file)
    history_manager = ModelEvaluationHistory(history_dir)
    return history_manager.add_evaluation_record(model_type, model_results, best_model_name)


def compare_and_select_best_model(model_type, config, test_data, device, history_file=None):
    """
    比较三种模型（基础模型、历史最佳模型和当前训练模型）并选择最佳模型
    
    Args:
        model_type: 模型类型，'tokenizer' 或 'predictor'
        config: 配置参数
        test_data: 测试数据
        device: 设备
        history_file: 历史记录文件路径，用于记录每天的模型性能，默认为None
        
    Returns:
        tuple: (最佳模型路径, 最佳模型损失, 模型比较结果字典)
    """
    try:
        models_to_compare = {}
        model_results = {}
        
        # 确定要比较的模型路径和名称
        if model_type == 'tokenizer':
            # 基础模型（远程模型）
            base_model_path = config.pretrained_tokenizer_path
            models_to_compare['base'] = {
                'path': base_model_path,
                'is_remote': True  # 基础模型从远程加载
            }
            
            # 历史最佳模型
            history_best_path = config.finetuned_tokenizer_path
            if os.path.exists(history_best_path):
                models_to_compare['history'] = {
                    'path': history_best_path,
                    'is_remote': False  # 历史最佳模型从本地加载
                }
            
            # 当前训练模型
            current_model_path = os.path.join(config.save_path, config.tokenizer_save_folder_name, 'checkpoints/best_model')
            if os.path.exists(current_model_path):
                models_to_compare['current'] = {
                    'path': current_model_path,
                    'is_remote': False  # 当前模型从本地加载
                }
                
            # 评估每个模型
            for model_name, model_info in models_to_compare.items():
                try:
                    test_loss = evaluate_tokenizer_on_test_data(
                        model_info['path'], 
                        test_data, 
                        config, 
                        device, 
                        is_remote=model_info['is_remote']
                    )
                    model_results[model_name] = {
                        'path': model_info['path'],
                        'loss': test_loss,
                        'date': datetime.now().strftime('%Y-%m-%d'),
                        'is_remote': model_info['is_remote']
                    }
                    logger.info(f"{model_name} 分词模型在测试集上的损失: {test_loss:.4f}")
                except Exception as e:
                    logger.error(f"评估 {model_name} 分词模型时出错: {str(e)}")
        
        else:  # predictor
            # 获取分词器路径
            tokenizer_path = config.finetuned_tokenizer_path
            
            # 基础模型（远程模型）
            base_model_path = config.pretrained_predictor_path
            models_to_compare['base'] = {
                'path': base_model_path,
                'is_remote': True  # 基础模型从远程加载
            }
            
            # 历史最佳模型
            history_best_path = config.finetuned_predictor_path
            if os.path.exists(history_best_path):
                models_to_compare['history'] = {
                    'path': history_best_path,
                    'is_remote': False  # 历史最佳模型从本地加载
                }
            
            # 当前训练模型
            current_model_path = os.path.join(config.save_path, config.predictor_save_folder_name, 'checkpoints/best_model')
            if os.path.exists(current_model_path):
                models_to_compare['current'] = {
                    'path': current_model_path,
                    'is_remote': False  # 当前模型从本地加载
                }
                
            # 评估每个模型
            for model_name, model_info in models_to_compare.items():
                try:
                    test_loss = evaluate_predictor_on_test_data(
                        model_info['path'], 
                        tokenizer_path, 
                        test_data, 
                        config, 
                        device,
                        is_remote_predictor=model_info['is_remote']
                    )
                    model_results[model_name] = {
                        'path': model_info['path'],
                        'loss': test_loss,
                        'date': datetime.now().strftime('%Y-%m-%d'),
                        'is_remote': model_info['is_remote']
                    }
                    logger.info(f"{model_name} 预测模型在测试集上的损失: {test_loss:.4f}")
                except Exception as e:
                    logger.error(f"评估 {model_name} 预测模型时出错: {str(e)}")
        
        # 找出最佳模型
        best_model_name = None
        best_model_loss = float('inf')
        best_model_path = None
        
        for model_name, result in model_results.items():
            if result['loss'] < best_model_loss:
                best_model_loss = result['loss']
                best_model_name = model_name
                best_model_path = result['path']
        
        if best_model_name:
            logger.info(f"最佳{model_type}模型是 {best_model_name} 模型，损失: {best_model_loss:.4f}")
            
            # 如果有历史文件，保存评估结果
            if history_file:
                save_model_evaluation_history(history_file, model_type, model_results, best_model_name)
                
            return best_model_path, best_model_loss, model_results
        else:
            logger.error(f"未能找到有效的{model_type}模型")
            return None, float('inf'), model_results
            
    except Exception as e:
        logger.error(f"比较模型过程中出错: {str(e)}")
        logger.error(traceback.format_exc())
        return None, float('inf'), {}


def evaluate_models_during_training(epoch_idx, current_model_path, config, test_data, device, model_type, best_loss, save_path):
    """
    在训练过程中评估模型并选择最佳模型
    
    该函数实现了高效的模型评估策略，通过只在第一个epoch评估所有可用模型，
    后续epoch只评估当前模型，大大减少了评估时间。
    
    优化策略：
    - 第一个epoch时评估所有可用模型（基础模型、历史最佳模型、当前模型）
    - 后续epoch只评估当前模型，并与已知的最佳模型比较
    - 每次找到更好的模型时，更新配置中的最佳模型路径
    
    工作流程:
    1. 第一个epoch (epoch_idx=0):
       - 评估所有可用模型（基础模型、历史最佳模型、当前模型）
       - 选择损失最小的模型作为最佳模型
       - 将最佳模型复制到指定路径
       - 更新配置中的最佳模型路径
    
    2. 后续epoch (epoch_idx>0):
       - 只评估当前模型
       - 将当前模型与已知的最佳模型比较
       - 如果当前模型更好，则更新最佳模型
    
    Args:
        epoch_idx: 当前训练轮次，0表示第一个epoch
        current_model_path: 当前训练的模型路径
        config: 配置参数对象，必须包含相关模型路径
        test_data: 测试数据，用于评估模型性能
        device: 设备（CPU或GPU）
        model_type: 模型类型，'tokenizer' 或 'predictor'
        best_loss: 当前已知的最佳损失值
        save_path: 保存最佳模型的路径
        
    Returns:
        tuple: (更新后的最佳损失值, 评估信息字典)
        评估信息字典包含:
            - epoch: 当前轮次
            - best_loss: 最佳损失
            - best_model_name: 最佳模型名称（base/history/current或epoch_X）
            - best_model_path: 最佳模型路径
            - evaluated_models: 所有评估的模型信息列表
    
    注意:
        - 该函数会修改config对象，更新his_best_tokenizer_path或his_best_predictor_path属性
        - 该函数会复制模型文件到指定路径，可能会覆盖已有文件
    """
    # 初始化评估信息
    eval_info = {
        'epoch': epoch_idx + 1,
        'best_loss': best_loss,
        'best_model_name': None,
        'best_model_path': None,
        'evaluated_models': []
    }
    
    try:
        # 第一个epoch时评估所有模型
        if epoch_idx == 0:
            logger.info(f"第一个epoch，评估所有可用的{model_type}模型...")
            best_model_path = None
            
            # 准备要评估的模型路径
            models_to_evaluate = {}
            
            if model_type == 'tokenizer':
                # 添加基础模型（远程模型）
                base_path = config.pretrained_tokenizer_path
                models_to_evaluate['base'] = {
                    'path': base_path,
                    'is_remote': True  # 基础模型从远程加载
                }
                
                # 添加历史最佳模型（如果存在且不同于基础模型）
                if hasattr(config, 'his_best_tokenizer_path') and config.his_best_tokenizer_path:
                    if os.path.exists(config.his_best_tokenizer_path) and config.his_best_tokenizer_path != base_path:
                        models_to_evaluate['history'] = {
                            'path': config.his_best_tokenizer_path,
                            'is_remote': False  # 历史最佳模型从本地加载
                        }
                
                # 添加当前模型
                if current_model_path not in [info['path'] for info in models_to_evaluate.values()]:
                    models_to_evaluate['current'] = {
                        'path': current_model_path,
                        'is_remote': False  # 当前模型从本地加载
                    }
                
                # 评估所有模型
                for model_name, model_info in models_to_evaluate.items():
                    try:
                        test_loss = evaluate_tokenizer_on_test_data(
                            model_info['path'], 
                            test_data, 
                            config, 
                            device, 
                            is_remote=model_info['is_remote']
                        )
                        logger.info(f"{model_name} 分词模型在测试集上的损失: {test_loss:.4f}")
                        
                        # 记录评估结果
                        eval_info['evaluated_models'].append({
                            'name': model_name,
                            'path': model_info['path'],
                            'loss': test_loss,
                            'is_remote': model_info['is_remote']
                        })
                        
                        # 更新最佳模型
                        if test_loss < best_loss:
                            best_loss = test_loss
                            best_model_path = model_info['path']
                            eval_info['best_model_name'] = model_name
                            eval_info['best_model_path'] = model_info['path']
                            eval_info['best_loss'] = best_loss
                            logger.info(f"{model_name} 分词模型成为新的最佳模型，损失: {best_loss:.4f}")
                    except Exception as e:
                        logger.error(f"评估 {model_name} 分词模型时出错: {str(e)}")
                        eval_info['evaluated_models'].append({
                            'name': model_name,
                            'path': model_info['path'],
                            'loss': float('inf'),
                            'error': str(e)
                        })
                
                # 如果找到了最佳模型，复制到指定路径
                if best_model_path:
                    os.makedirs(save_path, exist_ok=True)
                    os.system(f"cp -r {best_model_path}/* {save_path}/")
                    logger.info(f"最佳分词模型已复制到 {save_path}")
                    
                    # 更新配置中的最佳模型路径
                    config.his_best_tokenizer_path = best_model_path
            
            else:  # predictor
                tokenizer_path = config.finetuned_tokenizer_path
                
                # 添加基础模型（远程模型）
                base_path = config.pretrained_predictor_path
                models_to_evaluate['base'] = {
                    'path': base_path,
                    'is_remote': True  # 基础模型从远程加载
                }
                
                # 添加历史最佳模型（如果存在且不同于基础模型）
                if hasattr(config, 'his_best_predictor_path') and config.his_best_predictor_path:
                    if os.path.exists(config.his_best_predictor_path) and config.his_best_predictor_path != base_path:
                        models_to_evaluate['history'] = {
                            'path': config.his_best_predictor_path,
                            'is_remote': False  # 历史最佳模型从本地加载
                        }
                
                # 添加当前模型
                if current_model_path not in [info['path'] for info in models_to_evaluate.values()]:
                    models_to_evaluate['current'] = {
                        'path': current_model_path,
                        'is_remote': False  # 当前模型从本地加载
                    }
                
                # 评估所有模型
                for model_name, model_info in models_to_evaluate.items():
                    try:
                        test_loss = evaluate_predictor_on_test_data(
                            model_info['path'], 
                            tokenizer_path, 
                            test_data, 
                            config, 
                            device,
                            is_remote_predictor=model_info['is_remote']
                        )
                        logger.info(f"{model_name} 预测模型在测试集上的损失: {test_loss:.4f}")
                        
                        # 记录评估结果
                        eval_info['evaluated_models'].append({
                            'name': model_name,
                            'path': model_info['path'],
                            'loss': test_loss,
                            'is_remote': model_info['is_remote']
                        })
                        
                        # 更新最佳模型
                        if test_loss < best_loss:
                            best_loss = test_loss
                            best_model_path = model_info['path']
                            eval_info['best_model_name'] = model_name
                            eval_info['best_model_path'] = model_info['path']
                            eval_info['best_loss'] = best_loss
                            logger.info(f"{model_name} 预测模型成为新的最佳模型，损失: {best_loss:.4f}")
                    except Exception as e:
                        logger.error(f"评估 {model_name} 预测模型时出错: {str(e)}")
                        eval_info['evaluated_models'].append({
                            'name': model_name,
                            'path': model_info['path'],
                            'loss': float('inf'),
                            'error': str(e)
                        })
                
                # 如果找到了最佳模型，复制到指定路径
                if best_model_path:
                    os.makedirs(save_path, exist_ok=True)
                    os.system(f"cp -r {best_model_path}/* {save_path}/")
                    logger.info(f"最佳预测模型已复制到 {save_path}")
                    
                    # 更新配置中的最佳模型路径
                    config.his_best_predictor_path = best_model_path
        
        # 后续epoch只评估当前模型
        else:
            logger.info(f"评估当前{model_type}模型...")
            
            if model_type == 'tokenizer':
                current_test_loss = evaluate_tokenizer_on_test_data(current_model_path, test_data, config, device)
            else:  # predictor
                tokenizer_path = config.finetuned_tokenizer_path
                current_test_loss = evaluate_predictor_on_test_data(current_model_path, tokenizer_path, test_data, config, device)
            
            # 记录评估结果
            current_model_name = f"epoch_{epoch_idx + 1}"
            eval_info['evaluated_models'].append({
                'name': current_model_name,
                'path': current_model_path,
                'loss': current_test_loss,
                'is_remote': False
            })
            
            # 记录测试损失
            logger.info(f"当前{model_type}模型在测试集上的损失: {current_test_loss:.4f}")
            logger.info(f"历史最佳{model_type}模型在测试集上的损失: {best_loss:.4f}")
            
            # 如果当前模型更好，则更新最佳模型
            if current_test_loss < best_loss:
                best_loss = current_test_loss
                os.makedirs(save_path, exist_ok=True)
                os.system(f"cp -r {current_model_path}/* {save_path}/")
                logger.info(f"当前{model_type}模型是新的最佳模型，已保存到 {save_path}")
                
                # 更新评估信息
                eval_info['best_loss'] = best_loss
                eval_info['best_model_name'] = current_model_name
                eval_info['best_model_path'] = current_model_path
                
                # 更新配置中的最佳模型路径
                if model_type == 'tokenizer':
                    config.his_best_tokenizer_path = current_model_path
                else:
                    config.his_best_predictor_path = current_model_path
            else:
                logger.info(f"当前{model_type}模型未能超过历史最佳模型，保持原有最佳模型")
        
        return best_loss, eval_info
    except Exception as e:
        logger.error(f"评估模型过程中出错: {str(e)}")
        logger.error(traceback.format_exc())
        eval_info['error'] = str(e)
        return best_loss, eval_info


def update_best_model_paths(config, best_model_dir=None):
    """
    更新历史最佳模型路径
    
    该函数在训练结束后被调用，用于将当前训练过程中找到的最佳模型保存到历史模型目录中。
    这样，下次训练时可以直接加载这些历史最佳模型进行比较，而不需要重新评估所有模型。
    
    工作流程:
    1. 从配置中获取最佳模型路径（his_best_tokenizer_path 和 his_best_predictor_path）
    2. 将这些模型复制到指定的历史模型目录中
    3. 保存模型信息到JSON文件，包括路径、来源和日期
    
    Args:
        config: 配置参数对象，必须包含 his_best_tokenizer_path 和 his_best_predictor_path 属性
        best_model_dir: 最佳模型存储目录，默认为None。如果为None，则使用config.model_history_dir
        
    Returns:
        tuple: (更新状态, 分词模型路径, 预测模型路径)
            - 更新状态: 布尔值，表示更新是否成功
            - 分词模型路径: 更新后的分词模型路径，如果更新失败则为None
            - 预测模型路径: 更新后的预测模型路径，如果更新失败则为None
    
    注意:
        - 该函数会创建历史模型目录（如果不存在）
        - 如果目标目录已存在，会覆盖其中的文件
        - 函数会保存一个JSON文件，记录模型信息和更新日期
    """
    try:
        # 如果没有指定最佳模型目录，使用配置中的模型历史记录目录
        if not best_model_dir and hasattr(config, 'model_history_dir'):
            best_model_dir = config.model_history_dir
        
        # 如果仍然没有目录，使用默认目录
        if not best_model_dir:
            best_model_dir = "./model_history"
        
        # 创建最佳模型目录
        os.makedirs(best_model_dir, exist_ok=True)
        
        # 获取当前最佳模型路径
        tokenizer_path = None
        predictor_path = None
        
        # 分词模型
        if hasattr(config, 'his_best_tokenizer_path') and config.his_best_tokenizer_path:
            tokenizer_src = config.his_best_tokenizer_path
            tokenizer_dst = os.path.join(best_model_dir, 'best_tokenizer')
            if os.path.exists(tokenizer_src):
                # 创建目标目录
                os.makedirs(tokenizer_dst, exist_ok=True)
                # 复制模型文件
                os.system(f"cp -r {tokenizer_src}/* {tokenizer_dst}/")
                logger.info(f"已更新历史最佳分词模型: {tokenizer_src} -> {tokenizer_dst}")
                tokenizer_path = tokenizer_dst
        
        # 预测模型
        if hasattr(config, 'his_best_predictor_path') and config.his_best_predictor_path:
            predictor_src = config.his_best_predictor_path
            predictor_dst = os.path.join(best_model_dir, 'best_predictor')
            if os.path.exists(predictor_src):
                # 创建目标目录
                os.makedirs(predictor_dst, exist_ok=True)
                # 复制模型文件
                os.system(f"cp -r {predictor_src}/* {predictor_dst}/")
                logger.info(f"已更新历史最佳预测模型: {predictor_src} -> {predictor_dst}")
                predictor_path = predictor_dst
        
        # 保存最佳模型信息
        best_model_info = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'tokenizer': {
                'path': tokenizer_path,
                'source': config.his_best_tokenizer_path if hasattr(config, 'his_best_tokenizer_path') else None
            },
            'predictor': {
                'path': predictor_path,
                'source': config.his_best_predictor_path if hasattr(config, 'his_best_predictor_path') else None
            }
        }
        
        # 保存最佳模型信息到JSON文件
        info_file = os.path.join(best_model_dir, 'best_model_info.json')
        with open(info_file, 'w') as f:
            json.dump(best_model_info, f, indent=4)
        
        logger.info(f"已保存最佳模型信息: {info_file}")
        return True, tokenizer_path, predictor_path
    
    except Exception as e:
        logger.error(f"更新历史最佳模型路径时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return False, None, None


def update_models_after_training(config, test_data, device, history_dir=None):
    """
    在每天训练结束后更新模型
    
    Args:
        config: 配置参数
        test_data: 测试数据
        device: 设备
        history_dir: 历史记录目录，默认为None
        
    Returns:
        tuple: (更新状态, 分词模型路径, 预测模型路径)
    """
    try:
        logger.info("开始训练后模型更新流程...")
        results = {}
        
        # 初始化历史记录管理器
        history_manager = None
        history_file = None
        if history_dir:
            history_manager = ModelEvaluationHistory(history_dir)
            history_file = os.path.join(history_dir, 'model_history.json')
        
        # 1. 评估分词模型
        logger.info("评估分词模型...")
        tokenizer_path, tokenizer_loss, tokenizer_results = compare_and_select_best_model(
            model_type='tokenizer',
            config=config,
            test_data=test_data,
            device=device,
            history_file=history_file
        )
        
        if tokenizer_path:
            # 更新配置中的分词模型路径
            config.finetuned_tokenizer_path = tokenizer_path
            logger.info(f"更新配置中的分词模型路径: {tokenizer_path}")
            results['tokenizer'] = {
                'path': tokenizer_path,
                'loss': tokenizer_loss,
                'results': tokenizer_results
            }
            
            # 找出最佳模型名称
            best_tokenizer_name = None
            for name, result in tokenizer_results.items():
                if result['path'] == tokenizer_path:
                    best_tokenizer_name = name
                    break
            
            # 如果有历史记录管理器，直接使用它添加记录
            if history_manager and best_tokenizer_name:
                history_manager.add_evaluation_record('tokenizer', tokenizer_results, best_tokenizer_name)
        else:
            logger.warning("未能找到有效的分词模型，保持原有配置")
        
        # 2. 评估预测模型
        logger.info("评估预测模型...")
        predictor_path, predictor_loss, predictor_results = compare_and_select_best_model(
            model_type='predictor',
            config=config,
            test_data=test_data,
            device=device,
            history_file=history_file
        )
        
        if predictor_path:
            # 更新配置中的预测模型路径
            config.finetuned_predictor_path = predictor_path
            logger.info(f"更新配置中的预测模型路径: {predictor_path}")
            results['predictor'] = {
                'path': predictor_path,
                'loss': predictor_loss,
                'results': predictor_results
            }
            
            # 找出最佳模型名称
            best_predictor_name = None
            for name, result in predictor_results.items():
                if result['path'] == predictor_path:
                    best_predictor_name = name
                    break
            
            # 如果有历史记录管理器，直接使用它添加记录
            if history_manager and best_predictor_name:
                history_manager.add_evaluation_record('predictor', predictor_results, best_predictor_name)
        else:
            logger.warning("未能找到有效的预测模型，保持原有配置")
        
        # 3. 保存当天的模型评估结果
        if history_dir:
            today = datetime.now().strftime('%Y-%m-%d')
            daily_summary_file = os.path.join(history_dir, f"model_summary_{today}.json")
            with open(daily_summary_file, 'w') as f:
                json.dump(results, f, indent=4)
            logger.info(f"保存当天模型评估摘要: {daily_summary_file}")
        
        logger.info("训练后模型更新流程完成")
        return True, tokenizer_path, predictor_path
    except Exception as e:
        logger.error(f"训练后模型更新过程中出错: {str(e)}")
        logger.error(traceback.format_exc())
        return False, None, None


def evaluate_models(tokenizer_paths, model_paths, test_data, config, device, save_paths=None):
    """
    评估多个模型并选择最佳模型
    
    Args:
        tokenizer_paths: 分词器路径字典，格式为 {"model_name": "path"}
        model_paths: 模型路径字典，格式为 {"model_name": "path"}
        test_data: 测试数据
        config: 配置参数
        device: 设备
        save_paths: 可选，保存路径字典，格式为 {"tokenizer": "path", "model": "path"}
        
    Returns:
        tuple: (最佳模型名称, 最佳验证损失)
    """
    try:
        logger.info("开始评估模型...")
        best_model_name = None
        best_val_loss = float('inf')
        
        for model_name in model_paths.keys():
            model_path = model_paths[model_name]
            tokenizer_path = tokenizer_paths[model_name]
            
            # 检查模型路径是否存在
            if not (os.path.exists(model_path) and os.path.exists(tokenizer_path)):
                logger.warning(f"{model_name} 模型或分词器路径不存在，跳过评估: {model_path}, {tokenizer_path}")
                continue
                
            logger.info(f"评估 {model_name} 模型: {model_path}")
            tokenizer = KronosTokenizer.from_pretrained(tokenizer_path, local_files_only=True)
            tokenizer.eval().to(device)
            
            model = Kronos.from_pretrained(model_path, local_files_only=True)
            model.eval().to(device)
            
            # 计算测试集上的损失
            val_loss = evaluate_model_on_test_data(model, tokenizer, test_data, config, device)
            logger.info(f"{model_name} 模型在测试集上的损失: {val_loss:.4f}")
            
            # 更新最佳模型
            if val_loss < best_val_loss:
                best_val_loss = val_loss
                best_model_name = model_name
                
                # 如果提供了保存路径且最佳模型不是当前模型，复制到最终路径
                if save_paths and model_name != "current":
                    # 复制模型和分词器到最终路径
                    logger.info(f"将 {model_name} 模型复制为最终模型")
                    os.system(f"cp -r {model_path}/* {save_paths['model']}/")
                    os.system(f"cp -r {tokenizer_path}/* {save_paths['tokenizer']}/")
        
        logger.info(f"模型评估完成，最佳模型: {best_model_name}，测试损失: {best_val_loss:.4f}")
        return best_model_name, best_val_loss
    except Exception as e:
        logger.error(f"模型评估过程中出错: {str(e)}")
        logger.error(traceback.format_exc())
        return None, float('inf')
