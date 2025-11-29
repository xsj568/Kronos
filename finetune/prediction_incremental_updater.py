#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
预测结果增量更新器
用于将每天的预测结果追加到固定的Excel文件中
"""

import os
import logging
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Optional, List, Tuple
import traceback
import requests
import time
import sys

# 添加utils路径以导入get_shanghai_time
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from utils.training_pipeline_utils import get_shanghai_time
except ImportError:
    # 如果导入失败，使用简单的实现
    def get_shanghai_time():
        from datetime import timezone
        return datetime.now(timezone(timedelta(hours=8)))

# 设置日志
logger = logging.getLogger('KronosPipeline')


class PredictionIncrementalUpdater:
    """
    预测结果增量更新器
    
    功能：
    1. 将每天的预测结果追加到固定的Excel文件
    2. 每行包含：日期、股票代码、当前价格、预测价格、预测涨跌幅等
    3. 自动管理文件，避免重复数据
    """
    
    def __init__(self, master_excel_path: str = "./outputs/predictions_master.xlsx"):
        """
        初始化增量更新器
        
        Args:
            master_excel_path: 主Excel文件路径
        """
        self.master_excel_path = master_excel_path
        self.master_dir = os.path.dirname(master_excel_path)
        
        # 确保目录存在
        if self.master_dir:
            os.makedirs(self.master_dir, exist_ok=True)
        
        # 初始化主Excel文件
        self._initialize_master_excel()
    
    def _initialize_master_excel(self):
        """初始化主Excel文件，如果不存在则创建"""
        if not os.path.exists(self.master_excel_path):
            # 创建空的DataFrame并保存
            df = pd.DataFrame(columns=[
                '生成日期',  # 何时生成的预测（今天）
                '最新日期',  # 股票的最新数据日期
                '预测日期',  # 预测的目标日期（最新日期的下一个交易日）
                '股票代码',
                # 预测值
                '预测开盘价',
                '预测最高价',
                '预测最低价',
                '预测收盘价',
                '预测成交量',
                # 真实值（初始为空，下次训练时回填）
                '真实开盘价',
                '真实最高价',
                '真实最低价',
                '真实收盘价',
                '真实成交量',
                # 预测涨跌幅
                '预测开盘涨跌幅(%)',
                '预测最高涨跌幅(%)',
                '预测最低涨跌幅(%)',
                '预测收盘涨跌幅(%)',
                '预测成交量涨跌幅(%)',
                # 预测准确性（下次计算）
                '开盘价误差(%)',
                '收盘价误差(%)'
            ])
            
            try:
                df.to_excel(self.master_excel_path, index=False, sheet_name='预测历史')
                logger.info(f"创建主预测Excel文件: {self.master_excel_path}")
            except Exception as e:
                logger.error(f"创建主Excel文件失败: {str(e)}")
                raise
    
    def append_daily_predictions(self, prediction_data: pd.DataFrame, prediction_date: Optional[str] = None):
        """
        追加每天的预测结果到主Excel文件
        
        Args:
            prediction_data: 预测数据DataFrame，包含详细的预测信息
            prediction_date: 预测生成日期，格式为'YYYY-MM-DD'，如果为None则使用当前日期
        
        Returns:
            bool: 是否成功追加
        """
        try:
            if prediction_date is None:
                prediction_date = datetime.now().strftime('%Y-%m-%d')
            
            logger.info(f"开始追加预测结果到主Excel文件: {prediction_date}")
            
            # 读取现有数据
            try:
                existing_df = pd.read_excel(self.master_excel_path, sheet_name='预测历史')
            except Exception as e:
                logger.warning(f"读取现有Excel数据失败，将创建新文件: {str(e)}")
                existing_df = pd.DataFrame()
            
            # 检查是否已经有今天生成的预测
            if not existing_df.empty and '生成日期' in existing_df.columns:
                # 删除今天生成的旧预测（如果存在）
                existing_df = existing_df[existing_df['生成日期'] != prediction_date]
                logger.info(f"删除已存在的 {prediction_date} 生成的预测数据")
            
            # 准备新的预测数据
            # 假设 prediction_data 包含详细的预测信息
            new_records = []
            
            for _, row in prediction_data.iterrows():
                record = {
                    '生成日期': prediction_date,  # 预测生成日期（今天）
                    '最新日期': row.get('last_data_date', ''),  # 股票的最新数据日期
                    '预测日期': row.get('day_1_date', ''),  # 预测目标日期（最新日期的下一个交易日）
                    '股票代码': row.get('stock_code', ''),
                    # 预测值
                    '预测开盘价': row.get('day_1_open', 0),
                    '预测最高价': row.get('day_1_high', 0),
                    '预测最低价': row.get('day_1_low', 0),
                    '预测收盘价': row.get('day_1_close', 0),
                    '预测成交量': row.get('day_1_volume', 0),
                    # 真实值（初始为空，等下次训练时回填）
                    '真实开盘价': None,
                    '真实最高价': None,
                    '真实最低价': None,
                    '真实收盘价': None,
                    '真实成交量': None,
                    # 预测涨跌幅
                    '预测开盘涨跌幅(%)': row.get('day_1_open_change_pct', 0),
                    '预测最高涨跌幅(%)': row.get('day_1_high_change_pct', 0),
                    '预测最低涨跌幅(%)': row.get('day_1_low_change_pct', 0),
                    '预测收盘涨跌幅(%)': row.get('day_1_close_change_pct', 0),
                    '预测成交量涨跌幅(%)': row.get('day_1_volume_change_pct', 0),
                    # 预测准确性（初始为空）
                    '开盘价误差(%)': None,
                    '收盘价误差(%)': None
                }
                new_records.append(record)
            
            # 创建新的预测DataFrame
            new_df = pd.DataFrame(new_records)
            
            # 合并数据
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            
            # 回填历史记录的真实数据（在保存前完成）
            logger.info("开始回填历史预测记录的真实数据...")
            updated_df = self.backfill_real_data(combined_df)
            
            # 如果回填了数据，使用更新后的DataFrame
            if updated_df is not None:
                combined_df = updated_df
            
            # 按生成日期、最新日期、预测日期降序排列（最新的在前面）
            if not combined_df.empty:
                sort_columns = ['生成日期', '最新日期', '预测日期']
                # 只按存在的列排序
                sort_columns = [col for col in sort_columns if col in combined_df.columns]
                if sort_columns:
                    combined_df = combined_df.sort_values(sort_columns, ascending=False).reset_index(drop=True)
            
            # 计算涨幅相关列（在保存前完成）
            logger.info("计算涨幅相关列...")
            combined_df = self._calculate_price_change_columns(combined_df)
            
            # 保存到Excel
            with pd.ExcelWriter(self.master_excel_path, engine='openpyxl') as writer:
                # 保存主预测历史
                combined_df.to_excel(writer, sheet_name='预测历史', index=False)
                
                # 创建摘要表
                self._create_summary_sheet(writer, combined_df)
                
                # 应用样式
                self._apply_excel_styles(writer)
            
            logger.info(f"✓ 成功追加 {len(new_records)} 条预测记录到主Excel文件")
            logger.info(f"✓ 主Excel文件路径: {self.master_excel_path}")
            logger.info(f"✓ 总记录数: {len(combined_df)}")
            
            return True
        
        except Exception as e:
            logger.error(f"追加预测结果失败: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    def _calculate_price_change_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        计算涨幅相关列
        
        添加的列：
        1. 预测最高价相对开盘价涨幅(%) = (预测最高价 - 预测开盘价) / 预测开盘价 * 100
        2. 真实最高价相对开盘价涨幅(%) = (真实最高价 - 真实开盘价) / 真实开盘价 * 100 (如果有真实数据)
        3. 涨幅误差(%) = 预测涨幅 - 真实涨幅
        
        Args:
            df: 数据DataFrame
            
        Returns:
            pd.DataFrame: 添加了涨幅列的DataFrame
        """
        try:
            # 检查必要的列是否存在
            if '预测开盘价' not in df.columns or '预测最高价' not in df.columns:
                logger.warning("缺少必要的预测价格列，跳过涨幅计算")
                return df
            
            # 1. 计算预测的最高价和预测的开盘价的涨幅
            df['预测最高价相对开盘价涨幅(%)'] = np.where(
                df['预测开盘价'] > 0,
                (df['预测最高价'] - df['预测开盘价']) / df['预测开盘价'] * 100,
                np.nan
            )
            
            # 2. 计算真实的最高价和真实的开盘价的涨幅
            has_real_open = '真实开盘价' in df.columns
            has_real_high = '真实最高价' in df.columns
            
            if has_real_open and has_real_high:
                # 如果有真实开盘价和真实最高价，直接计算
                df['真实最高价相对开盘价涨幅(%)'] = np.where(
                    (df['真实开盘价'].notna()) & (df['真实开盘价'] > 0),
                    (df['真实最高价'] - df['真实开盘价']) / df['真实开盘价'] * 100,
                    np.nan
                )
            else:
                logger.debug("未找到真实开盘价或真实最高价列，跳过真实涨幅计算")
                df['真实最高价相对开盘价涨幅(%)'] = np.nan
            
            # 3. 计算两个涨幅之间的误差
            df['涨幅误差(%)'] = np.where(
                (df['预测最高价相对开盘价涨幅(%)'].notna()) & 
                (df['真实最高价相对开盘价涨幅(%)'].notna()),
                df['预测最高价相对开盘价涨幅(%)'] - df['真实最高价相对开盘价涨幅(%)'],
                np.nan
            )
            
            # 移除不需要的辅助列（如果存在）
            columns_to_remove = ['最新日期收盘价', '真实开盘价缺失']
            for col in columns_to_remove:
                if col in df.columns:
                    df = df.drop(columns=[col])
            
            # 统计信息
            has_prediction = df['预测最高价相对开盘价涨幅(%)'].notna().sum()
            has_real = df['真实最高价相对开盘价涨幅(%)'].notna().sum()
            has_error = df['涨幅误差(%)'].notna().sum()
            
            logger.info(f"  预测涨幅计算: {has_prediction}/{len(df)} 条记录")
            logger.info(f"  真实涨幅计算: {has_real}/{len(df)} 条记录")
            logger.info(f"  涨幅误差计算: {has_error}/{len(df)} 条记录")
            
            if has_error > 0:
                logger.info(f"  涨幅误差统计: 平均={df['涨幅误差(%)'].mean():.4f}%, "
                          f"中位数={df['涨幅误差(%)'].median():.4f}%, "
                          f"标准差={df['涨幅误差(%)'].std():.4f}%")
            
            return df
            
        except Exception as e:
            logger.warning(f"计算涨幅列失败: {str(e)}")
            logger.debug(traceback.format_exc())
            return df
    
    def _create_summary_sheet(self, writer, data_df):
        """创建摘要表"""
        try:
            if data_df.empty:
                return
            
            # 按预测日期统计
            date_col = '预测日期' if '预测日期' in data_df.columns else '目标日期'
            if date_col in data_df.columns:
                summary_by_date = data_df.groupby(date_col).agg({
                    '股票代码': 'count',
                    '预测收盘涨跌幅(%)': ['mean', 'std', 'min', 'max']
                }).reset_index()
                
                summary_by_date.columns = [
                    '预测日期',
                    '股票数量',
                    '平均预测涨跌幅(%)',
                    '涨跌幅标准差',
                    '最小涨跌幅(%)',
                    '最大涨跌幅(%)'
                ]
                
                summary_by_date.to_excel(writer, sheet_name='日期摘要', index=False)
                logger.info("✓ 创建日期摘要表")
        
        except Exception as e:
            logger.warning(f"创建摘要表失败: {str(e)}")
    
    def _apply_excel_styles(self, writer):
        """应用Excel样式"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 获取预测历史工作表
            if '预测历史' in writer.sheets:
                ws = writer.sheets['预测历史']
                
                # 设置表头样式
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 设置列宽
                for col_idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                
                # 为涨跌幅列添加条件格式
                for row_idx in range(2, ws.max_row + 1):
                    for col_idx in range(14, 19):  # 涨跌幅列
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            if cell.value > 0:
                                cell.font = Font(color='FF0000', bold=True)  # 红色
                            elif cell.value < 0:
                                cell.font = Font(color='00B050', bold=True)  # 绿色
                            cell.number_format = '0.00'
            
            logger.info("✓ 应用Excel样式")
        
        except ImportError:
            logger.warning("未安装openpyxl库，跳过样式设置")
        except Exception as e:
            logger.warning(f"应用Excel样式失败: {str(e)}")
    
    def get_latest_predictions(self, num_days: int = 7) -> pd.DataFrame:
        """
        获取最近N天的预测记录
        
        Args:
            num_days: 天数
        
        Returns:
            pd.DataFrame: 最近N天的预测记录
        """
        try:
            df = pd.read_excel(self.master_excel_path, sheet_name='预测历史')
            
            # 获取唯一的预测日期
            unique_dates = df['预测日期'].unique()[:num_days]
            
            # 筛选最近的记录
            recent_df = df[df['预测日期'].isin(unique_dates)]
            
            return recent_df
        
        except Exception as e:
            logger.error(f"获取最近预测记录失败: {str(e)}")
            return pd.DataFrame()
    
    def get_real_stock_data(self, stock_code: str, target_date: str) -> Optional[Dict]:
        """
        从sina API获取指定股票在指定日期的真实数据
        
        Args:
            stock_code: 股票代码
            target_date: 目标日期，格式为'YYYY-MM-DD'
        
        Returns:
            dict: 包含真实数据的字典，格式为 {'open': float, 'high': float, 'low': float, 'close': float, 'volume': int}
                 如果获取失败则返回None
        """
        try:
            url = f"http://stock.finance.sina.com.cn/usstock/api/json_v2.php/US_MinKService.getDailyK?symbol={stock_code}&___qn=3n"
            response = requests.get(url, timeout=10)
            
            if response.status_code != 200:
                return None
            
            data_json = response.json()
            if not data_json:
                return None
            
            # 查找目标日期的数据
            target_date_obj = pd.to_datetime(target_date)
            for item in data_json:
                date_str = item.get('d', '')
                if not date_str or date_str.startswith('0000-00-00'):
                    continue
                
                try:
                    item_date = pd.to_datetime(date_str)
                    if item_date.date() == target_date_obj.date():
                        return {
                            'open': float(item.get('o', 0)),
                            'high': float(item.get('h', 0)),
                            'low': float(item.get('l', 0)),
                            'close': float(item.get('c', 0)),
                            'volume': int(item.get('v', 0))
                        }
                except (ValueError, TypeError):
                    continue
            
            return None
            
        except Exception as e:
            logger.debug(f"获取股票 {stock_code} 在 {target_date} 的真实数据失败: {str(e)}")
            return None
    
    def backfill_real_data(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """
        回填历史预测记录的真实数据
        
        Args:
            df: 预测数据DataFrame
        
        Returns:
            pd.DataFrame: 更新后的DataFrame，如果无需更新则返回None
        """
        try:
            if df.empty:
                return None
            
            # 获取当前日期（使用上海时间，用于判断预测日期是否已过去）
            shanghai_time = get_shanghai_time()
            today = shanghai_time.date()
            
            # 检查哪些记录需要回填
            updated_df = df.copy()
            
            # 找出需要回填的记录（预测日期已过去，但真实数据为空）
            mask = (
                (pd.to_datetime(updated_df['预测日期'], errors='coerce').dt.date < today) &
                (updated_df['真实开盘价'].isna() | updated_df['真实收盘价'].isna())
            )
            
            records_to_backfill = updated_df[mask]
            
            if records_to_backfill.empty:
                logger.info("  没有需要回填的历史记录")
                return None
            
            logger.info(f"  找到 {len(records_to_backfill)} 条需要回填的记录")
            
            # 按股票代码和预测日期分组，避免重复请求
            unique_stocks_dates = records_to_backfill[['股票代码', '预测日期']].drop_duplicates()
            
            success_count = 0
            fail_count = 0
            
            for _, row in unique_stocks_dates.iterrows():
                stock_code = row['股票代码']
                predict_date = row['预测日期']
                
                # 获取真实数据
                real_data = self.get_real_stock_data(stock_code, predict_date)
                
                if real_data:
                    # 更新所有匹配的记录
                    update_mask = (
                        (updated_df['股票代码'] == stock_code) &
                        (updated_df['预测日期'] == predict_date)
                    )
                    
                    updated_df.loc[update_mask, '真实开盘价'] = real_data['open']
                    updated_df.loc[update_mask, '真实最高价'] = real_data['high']
                    updated_df.loc[update_mask, '真实最低价'] = real_data['low']
                    updated_df.loc[update_mask, '真实收盘价'] = real_data['close']
                    updated_df.loc[update_mask, '真实成交量'] = real_data['volume']
                    
                    # 计算预测误差
                    for idx in updated_df[update_mask].index:
                        pred_open = updated_df.loc[idx, '预测开盘价']
                        pred_close = updated_df.loc[idx, '预测收盘价']
                        real_open = real_data['open']
                        real_close = real_data['close']
                        
                        if pd.notna(pred_open) and real_open > 0:
                            updated_df.loc[idx, '开盘价误差(%)'] = ((pred_open - real_open) / real_open) * 100
                        
                        if pd.notna(pred_close) and real_close > 0:
                            updated_df.loc[idx, '收盘价误差(%)'] = ((pred_close - real_close) / real_close) * 100
                    
                    success_count += 1
                else:
                    fail_count += 1
                
                # 避免请求过快
                time.sleep(0.1)
            
            logger.info(f"  回填完成: 成功 {success_count} 条, 失败 {fail_count} 条")
            
            return updated_df
            
        except Exception as e:
            logger.error(f"回填真实数据失败: {str(e)}")
            logger.error(traceback.format_exc())
            return None
    
    def export_to_csv(self, csv_path: Optional[str] = None):
        """
        导出预测历史到CSV文件
        
        Args:
            csv_path: CSV文件路径，如果为None则使用默认路径
        
        Returns:
            bool: 是否成功导出
        """
        try:
            if csv_path is None:
                csv_path = self.master_excel_path.replace('.xlsx', '.csv')
            
            df = pd.read_excel(self.master_excel_path, sheet_name='预测历史')
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            
            logger.info(f"✓ 成功导出预测历史到CSV: {csv_path}")
            return True
        
        except Exception as e:
            logger.error(f"导出CSV失败: {str(e)}")
            return False


def update_master_predictions(prediction_df: pd.DataFrame, 
                               master_excel_path: str = "./outputs/predictions_master.xlsx",
                               prediction_date: Optional[str] = None) -> bool:
    """
    便捷函数：更新主预测Excel文件
    
    Args:
        prediction_df: 预测数据DataFrame
        master_excel_path: 主Excel文件路径
        prediction_date: 预测日期
    
    Returns:
        bool: 是否成功更新
    """
    updater = PredictionIncrementalUpdater(master_excel_path)
    return updater.append_daily_predictions(prediction_df, prediction_date)


if __name__ == '__main__':
    # 测试增量更新器
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s | %(levelname)-8s | %(message)s'
    )
    
    logger.info("=== 测试预测增量更新器 ===")
    
    # 创建测试数据
    test_data = {
        'stock_code': ['AAPL', 'MSFT', 'GOOG'],
        'current_open': [150.0, 300.0, 2800.0],
        'current_high': [152.0, 305.0, 2850.0],
        'current_low': [149.0, 298.0, 2780.0],
        'current_close': [151.0, 302.0, 2820.0],
        'current_volume': [1000000, 800000, 500000],
        'day_1_date': ['2024-01-02', '2024-01-02', '2024-01-02'],
        'day_1_open': [152.0, 305.0, 2850.0],
        'day_1_high': [155.0, 310.0, 2900.0],
        'day_1_low': [150.0, 303.0, 2840.0],
        'day_1_close': [153.0, 308.0, 2870.0],
        'day_1_volume': [1100000, 850000, 520000],
        'day_1_open_change_pct': [1.32, 1.66, 1.77],
        'day_1_high_change_pct': [1.97, 1.64, 1.75],
        'day_1_low_change_pct': [0.67, 1.68, 2.15],
        'day_1_close_change_pct': [1.32, 1.99, 1.77],
        'day_1_volume_change_pct': [10.0, 6.25, 4.0]
    }
    
    test_df = pd.DataFrame(test_data)
    
    # 测试更新
    updater = PredictionIncrementalUpdater("./test_predictions_master.xlsx")
    success = updater.append_daily_predictions(test_df, "2024-01-01")
    
    if success:
        logger.info("✓ 测试成功")
        
        # 获取最近的预测
        recent = updater.get_latest_predictions(num_days=1)
        logger.info(f"最近的预测记录数: {len(recent)}")
        
        # 导出CSV
        updater.export_to_csv("./test_predictions_master.csv")
    else:
        logger.error("✗ 测试失败")

